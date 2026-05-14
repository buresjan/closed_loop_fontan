#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import scipy.io

try:
    import yaml
except ImportError:  # pragma: no cover - requirements include pyyaml.
    yaml = None


ROOT = Path(__file__).resolve().parents[2]
RAW_ZIP = ROOT / "data/raw/Aramburu_et_al_2024_Heliyon_e30404.zip"
RAW_DIR = ROOT / "data/raw/Aramburu_et_al_2024_Heliyon_e30404"
OUT_DIR = ROOT / "data/processed/aramburu_2024"

MMHG_TO_PA = 133.33
ML_TO_M3 = 1e-6

MEASUREMENT_VARIABLES = [
    ("t", "time", "time", "s", "s"),
    ("P_IVC", "ivc_pressure", "pressure", "mmHg", "Pa"),
    ("P_LPA", "lpa_pressure", "pressure", "mmHg", "Pa"),
    ("P_RPA", "rpa_pressure", "pressure", "mmHg", "Pa"),
    ("P_SVC", "svc_pressure", "pressure", "mmHg", "Pa"),
    ("P_arch", "aortic_arch_pressure", "pressure", "mmHg", "Pa"),
    ("P_ascAo", "ascending_aorta_pressure", "pressure", "mmHg", "Pa"),
    ("P_dAo", "descending_aorta_pressure", "pressure", "mmHg", "Pa"),
    ("P_fem_art", "femoral_artery_pressure", "pressure", "mmHg", "Pa"),
    ("P_ventricle", "ventricle_pressure", "pressure", "mmHg", "Pa"),
    ("P_wedge", "wedge_pressure", "pressure", "mmHg", "Pa"),
    ("Q_IVC", "ivc_flow", "flow", "ml/s", "m3/s"),
    ("Q_LPA", "lpa_flow", "flow", "ml/s", "m3/s"),
    ("Q_RPA", "rpa_flow", "flow", "ml/s", "m3/s"),
    ("Q_SVC", "svc_flow", "flow", "ml/s", "m3/s"),
    ("Q_ascAo", "ascending_aorta_flow", "flow", "ml/s", "m3/s"),
    ("Q_dAo", "descending_aorta_flow", "flow", "ml/s", "m3/s"),
    ("V_ventricle", "ventricle_volume", "volume", "ml", "m3"),
]

SIM_CASES = [
    ("01_aorta", "Nektar_Simulations/01_Aorta/simulation_data.mat", "M004art_aorta"),
    ("02_tcpc", "Nektar_Simulations/02_TCPC/simulation_data.mat", "M005art_Fcross"),
    ("03_aorta_tcpc", "Nektar_Simulations/03_Aorta-TCPC/simulation_data.mat", "M009art_coupled"),
    (
        "04_aorta_tcpc_closedloop",
        "Nektar_Simulations/04_Aorta-TCPC_closedloop/simulation_data.mat",
        "M009art_closedloop_03",
    ),
]

SELECTED_NEKTAR_FIELDS = [
    ("P", "pressure_pa"),
    ("Pe", "elastic_pressure_pa"),
    ("Q", "flow_m3_s"),
    ("U", "velocity_m_s"),
    ("A", "area_m2"),
    ("Pavg", "pressure_avg_pa"),
    ("Qavg", "flow_avg_m3_s"),
]

LAST_CYCLE_SLICE = slice(34279, 35137)


def ensure_raw_data() -> None:
    if RAW_DIR.exists():
        return
    if not RAW_ZIP.exists():
        raise FileNotFoundError(
            f"Expected raw archive at {RAW_ZIP}. Place the Aramburu zip there first."
        )
    with zipfile.ZipFile(RAW_ZIP) as zf:
        zf.extractall(RAW_ZIP.parent)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_yaml(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if yaml is not None:
        path.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    else:
        path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def clean_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("%", "percent")
    text = text.replace("/", "_per_")
    text = text.replace("-", "_")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "column"


def unique_columns(columns: list[Any]) -> list[str]:
    out: list[str] = []
    seen: dict[str, int] = {}
    for col in columns:
        base = clean_name(col)
        seen[base] = seen.get(base, 0) + 1
        out.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return out


def sheet_filename(name: str) -> str:
    return clean_name(name).replace("_", "-") + ".csv"


def write_df(path: Path, df: pd.DataFrame, **kwargs: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, **kwargs)


def convert_measurements() -> list[dict[str, Any]]:
    mat = scipy.io.loadmat(RAW_DIR / "measurements.mat", squeeze_me=True)
    clinical: dict[str, np.ndarray] = {}
    si: dict[str, np.ndarray] = {}
    variable_docs: list[dict[str, Any]] = []

    for source, canonical, kind, clinical_unit, si_unit in MEASUREMENT_VARIABLES:
        values = np.asarray(mat[source]).reshape(-1)
        if kind == "pressure":
            clinical[f"{canonical}_mmHg"] = values
            si[f"{canonical}_pa"] = values * MMHG_TO_PA
            conversion = "Pa = mmHg * 133.33"
        elif kind == "flow":
            clinical[f"{canonical}_ml_s"] = values
            si[f"{canonical}_m3_s"] = values * ML_TO_M3
            conversion = "m3/s = ml/s * 1e-6"
        elif kind == "volume":
            clinical[f"{canonical}_ml"] = values
            si[f"{canonical}_m3"] = values * ML_TO_M3
            conversion = "m3 = ml * 1e-6"
        else:
            clinical[f"{canonical}_s"] = values
            si[f"{canonical}_s"] = values
            conversion = "identity"

        variable_docs.append(
            {
                "source": source,
                "canonical": canonical,
                "quantity": kind,
                "source_unit": clinical_unit,
                "si_unit": si_unit,
                "conversion": conversion,
            }
        )

    write_df(OUT_DIR / "measurements_clinical.csv", pd.DataFrame(clinical))
    write_df(OUT_DIR / "measurements.csv", pd.DataFrame(si))
    write_df(OUT_DIR / "comparison/measurements_last_cycle_clinical.csv", pd.DataFrame(clinical))
    return variable_docs


def workbook_sheet_to_csvs(workbook: Path, output_dir: Path) -> None:
    sheets = pd.read_excel(workbook, sheet_name=None, engine="openpyxl")
    for sheet, df in sheets.items():
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        df.columns = unique_columns(list(df.columns))
        write_df(output_dir / sheet_filename(sheet), df)


def cell_value(ws: Any, row: int, col: int) -> Any:
    value = ws.cell(row=row, column=col).value
    if isinstance(value, str):
        return value.strip()
    return value


def convert_geometry_sheet(ws: Any, output: Path) -> None:
    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        name = cell_value(ws, r, 1)
        if not name:
            continue
        rows.append(
            {
                "segment_name": name,
                "domain_id": cell_value(ws, r, 2),
                "node_1": cell_value(ws, r, 3),
                "node_2": cell_value(ws, r, 4),
                "length_m": cell_value(ws, r, 5),
                "radius_in_m": cell_value(ws, r, 6),
                "radius_out_m": cell_value(ws, r, 7),
            }
        )
    write_df(output, pd.DataFrame(rows))


def convert_time_table(ws: Any, output: Path, columns: list[str], start_row: int = 2) -> None:
    rows: list[dict[str, Any]] = []
    for r in range(start_row, ws.max_row + 1):
        values = [cell_value(ws, r, i + 1) for i in range(len(columns))]
        if values[0] is None:
            continue
        rows.append(dict(zip(columns, values)))
    write_df(output, pd.DataFrame(rows))


def convert_coupling_sheet(ws: Any, prefix: str, output_dir: Path) -> None:
    params: list[dict[str, Any]] = []
    for r in range(1, 6):
        label = cell_value(ws, r, 1)
        if not label:
            continue
        params.append(
            {
                "source_label": label,
                "value_source_units": cell_value(ws, r, 2),
                "value_si": cell_value(ws, r, 3),
            }
        )
    write_df(output_dir / f"{prefix}_parameters.csv", pd.DataFrame(params))

    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        t = cell_value(ws, r, 5)
        if t is None:
            continue
        rows.append(
            {
                "time_s": t,
                "measured_pressure_mmHg": cell_value(ws, r, 6),
                "simulated_pressure_mmHg": cell_value(ws, r, 7),
                "simulated_flow_ml_s": cell_value(ws, r, 8),
            }
        )
    write_df(output_dir / f"{prefix}_waveforms.csv", pd.DataFrame(rows))


def convert_data_workbook() -> None:
    wb = openpyxl.load_workbook(RAW_DIR / "data.xlsx", data_only=True, read_only=True)
    out = OUT_DIR / "model_inputs"

    convert_geometry_sheet(wb["Aorta"], out / "aorta_geometry.csv")
    convert_time_table(
        wb["Aorta waves"],
        out / "aorta_waves_clinical.csv",
        ["time_s", "ascending_aorta_flow_ml_s", "ascending_aorta_pressure_mmHg"],
    )
    convert_geometry_sheet(wb["Fontan cross"], out / "fontan_cross_geometry.csv")
    convert_time_table(
        wb["Fontan cross inflows"],
        out / "fontan_cross_inflows_clinical.csv",
        ["time_s", "ivc_flow_ml_s", "svc_flow_ml_s"],
    )
    convert_coupling_sheet(wb["Coupling_dAo_IVC"], "coupling_dao_ivc", out)
    convert_coupling_sheet(wb["Coupling_supAo_SVC"], "coupling_supao_svc", out)

    closed_loop_cells: list[dict[str, Any]] = []
    ws = wb["Closed-loop"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                closed_loop_cells.append(
                    {"row": cell.row, "column": cell.column, "value": cell.value}
                )
    write_df(out / "closed_loop_cells.csv", pd.DataFrame(closed_loop_cells))


def load_sim_struct(path: Path, top_field: str) -> np.ndarray:
    loaded = scipy.io.loadmat(
        path,
        squeeze_me=True,
        struct_as_record=False,
        variable_names=["data"],
    )["data"]
    return np.atleast_1d(getattr(loaded, top_field))


def domain_to_dataframe(domain: Any) -> pd.DataFrame:
    base = np.asarray(getattr(domain, "P"))
    n_time, n_positions = base.shape
    fs = float(getattr(domain, "fs"))
    time = np.arange(n_time, dtype=float) / fs
    distances = np.asarray(getattr(domain, "distances"), dtype=float).reshape(-1)

    parts: list[pd.DataFrame] = []
    for pos in range(n_positions):
        data: dict[str, Any] = {
            "time_s": time,
            "position_index": pos + 1,
            "distance_m": float(distances[pos]) if pos < len(distances) else np.nan,
        }
        for source, target in SELECTED_NEKTAR_FIELDS:
            if hasattr(domain, source):
                values = np.asarray(getattr(domain, source))
                if values.ndim == 2 and pos < values.shape[1]:
                    data[target] = values[:, pos]
        parts.append(pd.DataFrame(data))
    return pd.concat(parts, ignore_index=True)


def convert_nektar_data() -> None:
    for case_name, rel_path, top_field in SIM_CASES:
        case_dir = OUT_DIR / "nektar_1d" / case_name
        domains = load_sim_struct(RAW_DIR / rel_path, top_field)
        manifest_rows: list[dict[str, Any]] = []
        for domain in domains:
            domain_no = int(getattr(domain, "domain_no"))
            output_name = f"domain_{domain_no:02d}.csv.gz"
            df = domain_to_dataframe(domain)
            write_df(case_dir / output_name, df, compression="gzip")
            manifest_rows.append(
                {
                    "domain_no": domain_no,
                    "fs_hz": float(getattr(domain, "fs")),
                    "distances_m": json.dumps(
                        [float(x) for x in np.asarray(getattr(domain, "distances")).reshape(-1)]
                    ),
                    "output_file": output_name,
                    "columns": json.dumps(list(df.columns)),
                }
            )
        write_df(case_dir / "domain_manifest.csv", pd.DataFrame(manifest_rows))


def convert_fo_outputs() -> None:
    specs = [
        (
            "03_aorta_tcpc",
            RAW_DIR / "Nektar_Simulations/03_Aorta-TCPC/M009art_coupled.fo",
            [
                "time_s",
                "upper_body_arterial_pressure_mmHg",
                "upper_body_venous_pressure_mmHg",
                "lower_body_arterial_pressure_mmHg",
                "lower_body_venous_pressure_mmHg",
            ],
        ),
        (
            "04_aorta_tcpc_closedloop",
            RAW_DIR / "Nektar_Simulations/04_Aorta-TCPC_closedloop/M009art_closedloop_03.fo",
            [
                "time_s",
                "upper_body_arterial_pressure_mmHg",
                "upper_body_venous_pressure_mmHg",
                "lower_body_arterial_pressure_mmHg",
                "lower_body_venous_pressure_mmHg",
                "ventricle_volume_ml",
                "aortic_valve_flow_ml_s",
                "aortic_valve_state",
                "atrium_volume_ml",
                "mitral_valve_flow_ml_s",
                "mitral_valve_state",
                "total_blood_volume_ml",
                "pulmonary_inflow_qin5_ml_s",
                "aortic_valve_pressure_mmHg",
                "ventricle_pressure_mmHg",
                "atrium_pressure_mmHg",
            ],
        ),
    ]
    for case_name, path, columns in specs:
        data = np.loadtxt(path)
        if data.ndim == 1:
            data = data.reshape(1, -1)
        if data.shape[1] != len(columns):
            columns = columns[: data.shape[1]] + [
                f"fo_col_{i:02d}" for i in range(len(columns) + 1, data.shape[1] + 1)
            ]
        write_df(OUT_DIR / "fo_outputs" / f"{case_name}.csv", pd.DataFrame(data, columns=columns))


def comparison_from_case(case_name: str, rel_path: str, top_field: str, fo_path: Path | None = None) -> None:
    domains = load_sim_struct(RAW_DIR / rel_path, top_field)
    by_domain = {int(getattr(domain, "domain_no")): domain for domain in domains}
    measurements = pd.read_csv(OUT_DIR / "measurements_clinical.csv")
    rows = {"time_s": measurements["time_s"].to_numpy()}

    def add_pressure(name: str, domain_no: int, position: int) -> None:
        rows[f"{name}_pressure_mmHg"] = (
            np.asarray(by_domain[domain_no].P)[LAST_CYCLE_SLICE, position - 1] / MMHG_TO_PA
        )

    def add_flow(name: str, domain_no: int, position: int) -> None:
        rows[f"{name}_flow_ml_s"] = (
            np.asarray(by_domain[domain_no].Q)[LAST_CYCLE_SLICE, position - 1] / ML_TO_M3
        )

    add_pressure("ascending_aorta", 1, 1)
    add_flow("ascending_aorta", 1, 1)
    add_pressure("aortic_arch", 1, 3)
    add_pressure("descending_aorta", 2, 3)
    add_flow("descending_aorta", 2, 3)
    add_pressure("ivc", 5, 1)
    add_flow("ivc", 5, 1)
    add_pressure("svc", 9, 1)
    add_flow("svc", 9, 1)
    add_pressure("rpa", 6, 3)
    add_flow("rpa", 6, 3)
    add_pressure("lpa", 8, 3)
    add_flow("lpa", 8, 3)

    if fo_path is not None:
        fo = np.loadtxt(fo_path)
        rows["ventricle_volume_ml"] = fo[LAST_CYCLE_SLICE, 5]
        rows["mitral_valve_flow_ml_s"] = fo[LAST_CYCLE_SLICE, 9]
        rows["pulmonary_inflow_qin5_ml_s"] = fo[LAST_CYCLE_SLICE, 12]
        rows["ventricle_pressure_mmHg"] = fo[LAST_CYCLE_SLICE, 14]
        rows["atrium_pressure_mmHg"] = fo[LAST_CYCLE_SLICE, 15]
        rows["atrium_volume_ml"] = fo[LAST_CYCLE_SLICE, 8]

    write_df(OUT_DIR / "comparison" / f"{case_name}_last_cycle_clinical.csv", pd.DataFrame(rows))


def convert_comparison_tables() -> None:
    comparison_from_case(
        "03_aorta_tcpc_1d",
        "Nektar_Simulations/03_Aorta-TCPC/simulation_data.mat",
        "M009art_coupled",
    )
    comparison_from_case(
        "04_aorta_tcpc_closedloop_1d",
        "Nektar_Simulations/04_Aorta-TCPC_closedloop/simulation_data.mat",
        "M009art_closedloop_03",
        RAW_DIR / "Nektar_Simulations/04_Aorta-TCPC_closedloop/M009art_closedloop_03.fo",
    )


def source_manifest() -> dict[str, Any]:
    files = []
    for path in sorted(RAW_DIR.rglob("*")):
        if not path.is_file():
            continue
        files.append(
            {
                "path": str(path.relative_to(RAW_DIR)),
                "size_bytes": path.stat().st_size,
                "sha256": sha256(path),
            }
        )
    archive = None
    if RAW_ZIP.exists():
        archive = {
            "path": str(RAW_ZIP.relative_to(ROOT)),
            "size_bytes": RAW_ZIP.stat().st_size,
            "sha256": sha256(RAW_ZIP),
        }
    return {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source": "Aramburu_et_al_2024_Heliyon_e30404.zip",
        "source_reference": "Aramburu et al. 2024, Heliyon e30404",
        "raw_archive": archive,
        "raw_extracted_root": str(RAW_DIR.relative_to(ROOT)),
        "raw_files": files,
        "data_policy": "Raw archive and extracted binaries are ignored by Git; processed standardized outputs are tracked.",
    }


def write_readme() -> None:
    text = """# Aramburu 2024 Processed Data

This directory contains standardized calibration/reference data derived from `Aramburu_et_al_2024_Heliyon_e30404.zip`.

Tracked outputs include:

- `measurements.csv`: canonical SI measurement waveforms.
- `measurements_clinical.csv`: comparison-friendly waveforms in mmHg, ml/s, and ml.
- `paper_results/`: CSV exports of `results.xlsx`.
- `model_inputs/`: curated geometry, waveform, coupling, and closed-loop input tables from `data.xlsx`.
- `nektar_1d/`: converted selected Nektar 1-D MATLAB fields as gzipped per-domain CSV files.
- `fo_outputs/`: converted `.fo` 0-D/coupled output files.
- `comparison/`: last-cycle clinical tables aligned with the MATLAB comparison scripts.
- `manifest.yaml`: source checksums and provenance.
- `variables.yaml`: canonical variable names, units, and conversions.

The raw archive and extracted raw files stay under `data/raw/` and are ignored by Git.
"""
    (OUT_DIR / "README.md").write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Standardize Aramburu et al. 2024 Fontan data.")
    parser.add_argument("--keep-existing", action="store_true", help="Do not delete existing processed output first.")
    args = parser.parse_args()

    ensure_raw_data()
    if OUT_DIR.exists() and not args.keep_existing:
        shutil.rmtree(OUT_DIR)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    variables = convert_measurements()
    workbook_sheet_to_csvs(RAW_DIR / "results.xlsx", OUT_DIR / "paper_results")
    convert_data_workbook()
    convert_nektar_data()
    convert_fo_outputs()
    convert_comparison_tables()
    write_yaml(
        OUT_DIR / "variables.yaml",
        {
            "measurements": variables,
            "nektar_selected_fields": [
                {"source": source, "canonical": canonical}
                for source, canonical in SELECTED_NEKTAR_FIELDS
            ],
        },
    )
    write_yaml(OUT_DIR / "manifest.yaml", source_manifest())
    write_readme()
    print(f"Wrote standardized Aramburu data to {OUT_DIR}")


if __name__ == "__main__":
    main()
